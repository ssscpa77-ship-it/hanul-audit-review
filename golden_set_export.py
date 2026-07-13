"""골든셋 회귀기준 xlsx 생성·Hanul DB 배포."""

from __future__ import annotations

from io import BytesIO

import golden_set_catalog as gsc

_FONT = "맑은 고딕"


def _join(t: tuple[str, ...]) -> str:
    return "; ".join(t) if t else ""


def build_golden_set_workbook() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "개요"
    ws0["A1"] = "4대 중점 골든셋 회귀기준 — FY2026"
    ws0["A1"].font = Font(name=_FONT, bold=True, size=14, color="1F3864")
    ws0["A2"] = "체크리스트 checklist_id 와 1:1 연동 (focus_checklist_catalog.golden_set_ids)"
    ws0["A3"] = "자가검증: python3 golden_set_regression.py"

    ws = wb.create_sheet("골든셋")
    headers = [
        "case_id",
        "company_type",
        "focus_issue_no",
        "checklist_ids",
        "scenario",
        "file_pattern",
        "expected_gap_type",
        "must_have_defects",
        "must_not_have_defects",
        "expected_tier1_min",
        "expected_tier1_max",
        "tier1_recall_target",
        "false_positive_max",
        "mock_sheet_code",
        "mock_sheet_title",
        "notes",
    ]
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = hdr_fill
        c.font = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    issue_colors = {0: "F2F2F2", 1: "D6E4F0", 2: "E2EFDA", 3: "FFF2CC", 4: "FCE4D6"}
    for i, c in enumerate(gsc.GOLDEN_SET_CASES, 2):
        vals = [
            c.case_id,
            c.company_type,
            c.focus_issue_no,
            _join(c.checklist_ids),
            c.scenario,
            c.file_pattern,
            c.expected_gap_type,
            _join(c.must_have_defects),
            _join(c.must_not_have_defects),
            c.expected_tier1_min,
            c.expected_tier1_max,
            c.tier1_recall_target,
            c.false_positive_max,
            c.mock_sheet_code,
            c.mock_sheet_title,
            c.notes,
        ]
        fill = PatternFill("solid", fgColor=issue_colors.get(c.focus_issue_no, "FFFFFF"))
        for j, v in enumerate(vals, 1):
            cell = ws.cell(i, j, v)
            cell.font = Font(name=_FONT, size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if j <= 3:
                cell.fill = fill
        ws.row_dimensions[i].height = 36

    widths = [14, 10, 8, 18, 36, 16, 12, 28, 28, 8, 8, 8, 8, 12, 24, 24]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("체크리스트연결")
    ws2.append(["checklist_id", "golden_set_ids", "issue_no", "checklist_item"])
    ws2["A1"].fill = hdr_fill
    ws2["B1"].fill = hdr_fill
    ws2["C1"].fill = hdr_fill
    ws2["D1"].fill = hdr_fill
    import focus_checklist_catalog as cat

    ri = 2
    for r in cat.LISTED_CHECKLIST_2026 + cat.UNLISTED_CHECKLIST_2026:
        if not r.golden_set_ids:
            continue
        ws2.cell(ri, 1, r.checklist_id)
        ws2.cell(ri, 2, _join(r.golden_set_ids))
        ws2.cell(ri, 3, r.issue_no)
        ws2.cell(ri, 4, r.checklist_item)
        ri += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deploy_to_hanul_db() -> str:
    from pathlib import Path

    import knowledge_base as kb
    import review_guidelines as rg

    root = kb.SOURCE_DIR
    if isinstance(root, str):
        from pathlib import Path
        root = Path(root)
    target = root / rg.GUIDELINES_DB_SUBDIR
    target.mkdir(parents=True, exist_ok=True)
    path = target / rg.TEMPLATE_FILES["golden_set"]
    path.write_bytes(build_golden_set_workbook())
    local = Path(__file__).resolve().parent / "data" / "templates" / rg.TEMPLATE_FILES["golden_set"]
    local.parent.mkdir(parents=True, exist_ok=True)
    local.write_bytes(path.read_bytes())
    return str(path)


if __name__ == "__main__":
    print(deploy_to_hanul_db())
