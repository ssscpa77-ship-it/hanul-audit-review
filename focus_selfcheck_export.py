"""4대중점 자가진단 리뷰노트(체크리스트 xlsx) 생성기.

「4대중점_심층검토_보완지침_2026-08-09.md」의 산출물 생성 스크립트.

사용법
------
    python3 focus_selfcheck_export.py                 # 기본 경로에 상장·비상장 2종 생성
    python3 focus_selfcheck_export.py --out DIR       # 출력 폴더 지정

생성물
------
    4대중점_자가진단_리뷰노트_상장_FY2026.xlsx
    4대중점_자가진단_리뷰노트_비상장_FY2026.xlsx

각 파일 구성
    ① 개요        — 목적·사용법·판정기준
    ② 이슈별_필수조서 — 중점이슈별로 반드시 존재해야 할 조서(누락 점검)
    ③ 자가진단_체크리스트 — 절차 단계별 진단표(핵심)
    ④ 범례        — 게이트·판정·미비유형 정의
"""

from __future__ import annotations

import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import focus_procedure_steps as fps

FONT = "Arial"
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_ISSUE_FILL = PatternFill("solid", fgColor="D9E2F3")
_INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
_G_FILL = {
    "G1": PatternFill("solid", fgColor="FCE4D6"),
    "G2": PatternFill("solid", fgColor="FFF2CC"),
    "G3": PatternFill("solid", fgColor="E2EFDA"),
    "G4": PatternFill("solid", fgColor="DEEAF6"),
}
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, row: int, headers: list[str], widths: list[int]) -> None:
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 34


def _load_catalog(is_listed: bool) -> dict[str, dict]:
    """체크리스트 항목 메타(근거·사례·조서코드)를 checklist_id로 색인."""
    try:
        import focus_checklist_catalog as fcc

        rows = fcc.rows_for_listed() if is_listed else fcc.rows_for_unlisted()
    except Exception as exc:  # noqa: BLE001
        print(f"[경고] focus_checklist_catalog 로드 실패({exc}) — 메타 없이 생성", file=sys.stderr)
        return {}
    out: dict[str, dict] = {}
    for r in rows:
        out[r.checklist_id] = {
            "issue_no": r.issue_no,
            "issue_title": r.issue_title,
            "checklist_item": r.checklist_item,
            "violation_type": r.violation_type,
            "case_source": r.case_source,
            "case_example": r.case_example,
            "basis": r.basis,
            "standard_paragraphs": getattr(r, "standard_paragraphs", ""),
            "audit_standard_ref": getattr(r, "audit_standard_ref", ""),
            "additional_case_refs": getattr(r, "additional_case_refs", ""),
            "qna_refs": getattr(r, "qna_refs", ""),
            "qc_checklist_ref": getattr(r, "qc_checklist_ref", ""),
            "related_sheet_codes": "; ".join(r.related_sheet_codes or ()),
            "materiality_note": r.materiality_note,
        }
    return out


def _sheet_overview(wb: Workbook, is_listed: bool, n_items: int, n_steps: int) -> None:
    ws = wb.create_sheet("개요", 0)
    src = "금융감독원" if is_listed else "한국공인회계사회"
    kind = "상장회사" if is_listed else "비상장회사"
    pub = "2026.6.19." if is_listed else "2026.6.29."
    lines = [
        (f"4대 중점 회계이슈 자가진단 리뷰노트 — {src}({kind})", True),
        ("", False),
        (f"출처: {src} 「2026년 재무제표에 대한 중점심사 회계이슈 사전예고」({pub}. 보도자료)", False),
        ("      Hanul DB / 4대 중점사항 감리대상 / 2026년 원문 및 회계위반(오류) 예시 전건 반영", False),
        ("", False),
        ("목적", True),
        ("  4대 중점 회계이슈에 대하여 ① 중요한 감사절차가 감사조서에 포함되어 있는지,", False),
        ("  ② 검토 결과가 적정한지, ③ 누락된 항목이 있는지, ④ 검토절차가 미흡한 것은 없는지를", False),
        ("  절차 단계 단위로 자가진단하기 위한 체크리스트입니다.", False),
        ("", False),
        ("구성", True),
        (f"  · 체크리스트 항목 {n_items}건을 원자적 감사절차 {n_steps}단계로 분해", False),
        ("  · 각 단계에 4단계 진단 게이트(G1 절차 존재성 → G2 증거 충족성 → G3 절차 충실성 → G4 결론 적정성) 부여", False),
        ("  · 각 단계마다 동일 지적유형의 감리지적사례를 병기", False),
        ("", False),
        ("사용법", True),
        ("  1) [이슈별_필수조서] 시트에서 4개 중점이슈별 필수 조서의 존재 여부를 먼저 확인합니다.", False),
        ("     → 조서 자체가 없으면 그 이슈는 '조서누락'으로, 단계별 진단에 앞서 최우선 보완대상입니다.", False),
        ("  2) [자가진단_체크리스트] 시트에서 단계별로 판정(노란색 열)을 입력합니다.", False),
        ("     → 판정: 적합 / 보완필요 / 미비 / 해당없음 (드롭다운)", False),
        ("  3) '미비' 또는 '보완필요'로 판정된 단계는 [리뷰사항] 문안을 그대로 리뷰노트에 사용할 수 있습니다.", False),
        ("  4) 상위 게이트가 무너지면(G1 미비) 하위 게이트는 그 결과이므로, 근본원인 1건만 리뷰노트로 발행합니다.", False),
        ("", False),
        ("판정 기준", True),
        ("  적합(OK)      : 절차·증거·정량근거·결론이 모두 확인됨", False),
        ("  보완필요(WEAK): 절차는 수행되었으나 정량근거 또는 결론 기재가 불충분함", False),
        ("  미비(GAP)     : 절차 자체가 없거나, 감리지적 유형과 동일한 위험문구가 확인됨", False),
        ("  해당없음(N/A) : 회사 상황상 적용되지 않음 — 반드시 그 근거를 조서에 명시", False),
        ("", False),
        ("주의", True),
        ("  · '해당없음' 판정은 반드시 판단근거를 조서에 남겨야 합니다. 근거 없는 N/A는 그 자체가 지적사항입니다.", False),
        ("  · 본 체크리스트는 자가검토(hanul-002) 엔진 focus_selfcheck.py와 동일한 데이터(focus_procedure_steps.py)를", False),
        ("    사용하므로, 앱 진단결과와 본 시트의 항목·번호가 1:1로 대응합니다.", False),
        ("", False),
        ("작성: 한울회계법인 H-Pro · 2026-08-09", False),
    ]
    ws.column_dimensions["A"].width = 118
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT, bold=bold, size=12 if (bold and i == 1) else 10)
        c.alignment = Alignment(vertical="center")
    ws.sheet_view.showGridLines = False


def _sheet_coverage(wb: Workbook, is_listed: bool) -> None:
    ws = wb.create_sheet("이슈별_필수조서")
    headers = ["중점이슈", "이슈명", "필수 조서코드", "적용 판단 기준(trigger)",
               "조서 존재여부", "미존재 시 조치", "확인자", "확인일"]
    widths = [10, 30, 16, 42, 14, 62, 12, 12]
    ws.cell(row=1, column=1, value="① 중점이슈별 필수 조서 존재 점검 — 조서 자체의 누락을 가장 먼저 확인합니다")
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=12)
    _hdr(ws, 3, headers, widths)

    dv = DataValidation(type="list", formula1='"존재,미존재,해당없음"', allow_blank=True)
    ws.add_data_validation(dv)

    r = 4
    for cov in fps.coverage_for(is_listed):
        vals = [
            f"제{cov.issue_no}항",
            cov.issue_title,
            ", ".join(cov.required_sheet_codes),
            cov.trigger_hint,
            "",
            cov.absence_note,
            "",
            "",
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = _BORDER
            if j in (5, 7, 8):
                c.fill = _INPUT_FILL
        dv.add(ws.cell(row=r, column=5))
        ws.row_dimensions[r].height = 68
        r += 1
    ws.freeze_panes = "A4"


def _sheet_checklist(wb: Workbook, is_listed: bool, meta: dict[str, dict]) -> int:
    ws = wb.create_sheet("자가진단_체크리스트")
    headers = [
        "중점이슈", "이슈명", "체크ID", "체크항목", "단계ID", "순서",
        "게이트", "게이트명", "필수 감사절차(단계)",
        "조서에서 확인할 증거", "정량근거 요건", "적출 대상 위험문구",
        "미비유형", "판정", "발견사항 메모",
        "리뷰사항(문안)", "리뷰근거", "감리지적사례", "관련조서", "검토자", "검토일",
    ]
    widths = [9, 26, 9, 26, 12, 6, 7, 12, 40, 30, 22, 26, 11, 11, 24, 52, 34, 44, 12, 10, 11]
    ws.cell(row=1, column=1,
            value="② 절차 단계별 자가진단 — 노란색 열(판정·메모·검토자·검토일)에 입력하십시오")
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=12)
    _hdr(ws, 3, headers, widths)

    dv = DataValidation(
        type="list",
        formula1=f'"{fps.VERDICT_OK},{fps.VERDICT_WEAK},{fps.VERDICT_GAP},{fps.VERDICT_NA}"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)

    src_steps = fps.LISTED_STEPS if is_listed else fps.UNLISTED_STEPS
    r = 4
    for cid in sorted(src_steps, key=lambda x: (x[:2], x)):
        m = meta.get(cid, {})
        for step in src_steps[cid]:
            case_parts = []
            if step.case_ref and m.get("case_example"):
                case_parts.append(
                    f"[{step.case_ref} · {m.get('violation_type','')}] {m['case_example'][:150]}"
                )
            elif step.case_ref:
                case_parts.append(f"[{step.case_ref}]")
            if m.get("additional_case_refs"):
                case_parts.append(f"추가사례: {m['additional_case_refs'][:150]}")
            basis = m.get("basis", "")
            if m.get("standard_paragraphs"):
                basis = f"{basis} / {m['standard_paragraphs'][:110]}"
            if m.get("audit_standard_ref"):
                basis = f"{basis} / {m['audit_standard_ref'][:80]}"

            vals = [
                f"제{m.get('issue_no','')}항",
                m.get("issue_title", ""),
                cid,
                m.get("checklist_item", ""),
                step.step_id,
                step.seq,
                step.gate,
                step.gate_name,
                step.title,
                "; ".join(step.evidence_keys[:8]),
                "; ".join(step.depth_keys[:6]) if step.gate == "G3" else "",
                "; ".join(step.red_flags[:5]),
                step.failure_type,
                "",
                "",
                step.review_note,
                basis.strip(" /"),
                "\n".join(case_parts),
                m.get("related_sheet_codes", ""),
                "",
                "",
            ]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = _BORDER
                if j == 7:
                    c.fill = _G_FILL.get(step.gate, _G_FILL["G1"])
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if j in (14, 15, 20, 21):
                    c.fill = _INPUT_FILL
            dv.add(ws.cell(row=r, column=14))
            ws.row_dimensions[r].height = 58
            r += 1
    ws.freeze_panes = "F4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{r-1}"
    return r - 4


def _sheet_legend(wb: Workbook) -> None:
    ws = wb.create_sheet("범례")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 92
    _hdr(ws, 1, ["구분", "코드", "정의"], [14, 22, 92])
    rows = [
        ("진단 게이트", "G1 절차 존재성",
         "해당 감사절차를 수행한 흔적이 조서에 있는가. 없으면 '절차누락' — 가장 중대한 결손이며, "
         "하위 게이트 결손은 모두 이것의 결과이므로 근본원인 1건만 리뷰노트로 발행한다."),
        ("", "G2 증거 충족성",
         "절차의 근거가 되는 증빙·문서(계약서·조회서·평가보고서·등기부 등)가 확보·기재되었는가. "
         "없으면 '증빙부족'."),
        ("", "G3 절차 충실성",
         "정량근거(재계산·표본·모집단·대사·산정내역)가 있는가. 없으면 '절차미흡'. "
         "감리지적 유형과 동일한 위험문구가 실제로 확인되면 '미비'로 격상하고 억제하지 않는다."),
        ("", "G4 결론 적정성",
         "결론이 증거와 정합하며 상투문구('이상없음' 등)로 갈음되지 않았는가. 없으면 '결론부적정'."),
        ("판정", fps.VERDICT_OK, "절차·증거·정량근거·결론이 모두 확인됨."),
        ("", fps.VERDICT_WEAK, "절차는 수행되었으나 정량근거 또는 결론 기재가 불충분함."),
        ("", fps.VERDICT_GAP, "절차 자체가 없거나, 감리지적 유형과 동일한 위험문구가 확인됨."),
        ("", fps.VERDICT_NA,
         "회사 상황상 적용되지 않음. 반드시 판단근거를 조서에 명시할 것 — 근거 없는 N/A는 그 자체가 지적사항."),
        ("미비유형", "조서누락", "중점이슈 관련 조서 자체가 존재하지 않음(이슈 수준 결손)."),
        ("", "절차누락", "체크항목의 필수 감사절차를 수행한 흔적이 없음."),
        ("", "증빙부족", "절차는 있으나 근거 증빙·문서가 확인되지 않음."),
        ("", "절차미흡", "절차·증빙은 있으나 재계산·표본 등 정량적 뒷받침이 없음."),
        ("", "결론부적정", "결론이 상투문구뿐이거나 증거와 배치됨."),
        ("사례기호", "상장 A~L사",
         "금융감독원 「2026년 재무제표에 대한 중점심사 회계이슈 사전예고」(2026.6.19.) 회계위반 예시."),
        ("", "비상장 A~T사",
         "한국공인회계사회 「2026년 비상장회사 재무제표에 대한 2027년 중점심사 회계이슈 사전예고」"
         "(2026.6.29.) 회계오류 예시."),
        ("", "FSS·KICPA-연도-번호", "Hanul DB 감리지적사례(금감원·한공회 2020~2026년, PDF·hwp 480여 건) 참조번호."),
    ]
    for i, (a, b, c) in enumerate(rows, start=2):
        for j, v in enumerate((a, b, c), start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = Font(name=FONT, size=9, bold=(j <= 2 and bool(v)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER
        ws.row_dimensions[i].height = 34
    ws.freeze_panes = "A2"


def build(is_listed: bool, out_dir: str) -> str:
    meta = _load_catalog(is_listed)
    wb = Workbook()
    wb.remove(wb.active)
    src_steps = fps.LISTED_STEPS if is_listed else fps.UNLISTED_STEPS
    n_items = len(src_steps)
    n_steps = sum(len(v) for v in src_steps.values())
    _sheet_overview(wb, is_listed, n_items, n_steps)
    _sheet_coverage(wb, is_listed)
    written = _sheet_checklist(wb, is_listed, meta)
    _sheet_legend(wb)
    kind = "상장" if is_listed else "비상장"
    path = os.path.join(out_dir, f"4대중점_자가진단_리뷰노트_{kind}_FY2026.xlsx")
    os.makedirs(out_dir, exist_ok=True)
    wb.save(path)
    print(f"  생성: {os.path.basename(path)} — 체크항목 {n_items}건 · 절차단계 {written}행")
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description="4대중점 자가진단 리뷰노트 생성")
    ap.add_argument("--out", default="data/templates", help="출력 폴더")
    args = ap.parse_args()
    print("4대중점 자가진단 리뷰노트 생성")
    for listed in (True, False):
        build(listed, args.out)


if __name__ == "__main__":
    main()
