"""품질관리(QC) 전문가 관점 보조 점검."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter

from parser import ParsedDocument
import review_engine as re_engine
import output_formatter as out_fmt

_INDEPENDENCE_KW = ("독립성", "independence", "비감사", "이해상충")

_GC_NARRATIVE_RE = re.compile(
    r"계속기업|going\s*concern|유동성\s*위험|유동성\s*부족|"
    r"자본잠식|영업손실|지급\s*불능|"
    r"파산|청산|정리절차|회생|워크아웃|"
    r"부도|채무\s*불이행",
    re.I,
)
_GC_ASSUMPTION_RE = re.compile(
    r"계속기업\s*가정|going\s*concern\s*assumption|"
    r"가정\s*적정|가정에\s*대한|"
    r"중요한\s*의문|significant\s*doubt|"
    r"570|계속기업\s*불확실성",
    re.I,
)
_GC_UNCERTAINTY_RE = re.compile(
    r"계속기업\s*불확실|중요한\s*의문|significant\s*doubt|"
    r"완전\s*자본잠식|자본잠식|영업손실\s*지속|"
    r"유동성\s*위험|지급\s*불능|파산|청산|정리절차|부도|"
    r"상환.*어렵|회수.*불가|회수불능",
    re.I,
)
_ALLOWANCE_REVIEW_RE = re.compile(
    r"대손|기대신용손실|ECL|손실충당|연령분석|aging|회수가능|"
    r"충당금\s*산정|신용손실|손상기준",
    re.I,
)
_EQUITY_IMPAIR_RE = re.compile(
    r"지분법.*손상|투자주식.*손상|손상차손|회수가능가액|"
    r"순자산\s*가액|impairment|자산손상|지분법\s*적용",
    re.I,
)
_INSUFFICIENT_ALLOWANCE_RE = re.compile(
    r"충당금.*(?:과소|부족|불충분|미설정|미반영)|"
    r"대손.*(?:미반영|반영.*않)|추가\s*대손.*반영.*않|"
    r"ECL\s*0|충당\s*없음|산정\s*근거\s*없|"
    r"환입.*발생.*반영.*않|중요성.*낮아\s*반영.*않",
    re.I,
)
_LOGICAL_FLAW_RE = re.compile(
    r"자본잠식.*손상징후.*없|손상징후.*없.*자본잠식|"
    r"불일치|모순|근거\s*부족|산정\s*오류|과소계상|"
    r"상환.*어렵.*roll|roll.over.*가정|"
    r"충분히\s*반영.*않|미흡|부적절",
    re.I,
)
_FLAW_CELL_RULES: list[tuple[re.Pattern[str], str]] = [
    (
        re.compile(r"추가\s*대손.*반영.*않|대손.*반영.*않음", re.I),
        "계속기업·상환곤란에도 추가 대손충당금·손상반영을 하지 않음",
    ),
    (
        re.compile(r"자본잠식.*손상징후.*없|손상징후.*없.*판단", re.I),
        "자본잠식·불확실성과 손상징후 부재 판단이 논리적으로 상충",
    ),
    (
        re.compile(r"상환.*어렵.*roll|roll.over.*가정", re.I),
        "단기 상환·회수 곤란에도 장기 roll-over만 가정하여 충당·손상이 부족해 보임",
    ),
    (
        re.compile(r"환입.*발생.*반영.*않|중요성.*낮아\s*반영.*않", re.I),
        "대손충당금 환입·조정 필요로 보이나 중요성만으로 미반영",
    ),
    (
        re.compile(r"충당금.*(?:과소|부족|불충분|미설정)", re.I),
        "대손충당금 설정이 부족해 보임",
    ),
]
_REVIEW_DOCUMENTED_RE = re.compile(
    r"검토(?:함|하였|했|완료|결과|내역)|판단(?:함|하였|했|됨)|확인(?:함|하였|했|됨)|"
    r"분류(?:함|하였|했|됨)|적용(?:함|하였|했|됨)|수행(?:함|하였|했|됨)|"
    r"문서화|결론|전액\s*손상|손상처리|손상차손|손상반영|손상\s*발생|"
    r"회수가능|산정(?:함|하였|했|됨)|평가(?:함|하였|했|됨)|"
    r"근거|절차|조회|실사|입회|감사메모|감사절차|감사목적|"
    r"계획|고려|결정|분석|상계|청산|반영|적정|합리|"
    r"이상\s*없|문제\s*없|중단|전액|손상처리",
    re.I,
)
_SHEET_REVIEW_FRAME_RE = re.compile(
    r"감사(?:메모|절차|목적)|Procedure|Reference|Ⅲ\.",
    re.I,
)
_EXPLICIT_GAP_RE = re.compile(
    r"미반영|미검토|미수행|누락|부족|과소|없음|미확인|미기재|"
    r"추가\s*대손.*반영.*않|(?:대손|충당|손상|손상차손).{0,24}반영.*않|미흡",
    re.I,
)
_NO_ADDITIONAL_ALLOWANCE_CONCLUSION_RE = re.compile(
    r"추가\s*대손.*반영.*않|대손.*반영.*않음|"
    r"환입.*반영.*않|중요성.*낮아\s*반영.*않|금액적\s*중요성.*반영.*않",
    re.I,
)
_ALLOWANCE_OMISSION_JUSTIFICATION_RE = re.compile(
    r"회수\s*가능|회수가능|감평|상계|사용가치|공정가치|담보|재무제표\s*수령|"
    r"가결산|충분히|산정|근거|판단|금액적\s*중요성|중요성이\s*낮|"
    r"영업이익|손상징후.*없|코스트코|안정적",
    re.I,
)
_SCOPE_AUDIT_RE = re.compile(
    r"피감사|감사대상|본회사|보고기업|연결실체|당사의\s*계속",
    re.I,
)
_SCOPE_SUB_RE = re.compile(
    r"종속(?:회사|기업)?|관계(?:회사|기업)|지분법|공동(?:기업|통제)|"
    r"SAJ|관계기업",
    re.I,
)
_SCOPE_BORROWER_RE = re.compile(
    r"대여(?:처|금)|차입처|채무자|피대여|차주|회수가능|손상검토|"
    r"담보|부동산담보|장기대여",
    re.I,
)
_LOAN_SHEET_CODES = re.compile(r"^F(?:200|211|300|310)\b", re.I)
_SUB_SHEET_CODES = re.compile(r"^SAJ", re.I)

_SCOPE_LABELS = (
    "감사대상 회사",
    "종속·관계회사",
    "대여(차입)처·피대여 회사",
    "미확인(범위 명시 필요)",
)


@dataclass
class _GCMention:
    sheet_no: str
    sheet_title: str
    excel_ref: str
    excerpt: str
    scope: str


@dataclass
class _ImpairmentHit:
    sheet_no: str
    sheet_title: str
    excel_ref: str
    excerpt: str
    issue: str
    category: str  # allowance | equity | loan


def _first_sheet_meta(
    *groups: list[Any],
    default_no: str = "-",
    default_title: str = "-",
) -> tuple[str, str]:
    """리스트 첫 항목의 (조서번호, 제목). 비어 있으면 기본값."""
    for group in groups:
        if not group:
            continue
        item = group[0]
        no = str(getattr(item, "sheet_no", "") or default_no)
        title = str(getattr(item, "sheet_title", "") or default_title)
        return _short_sheet(no), title
    return default_no, default_title


def run_qc_checks(doc: ParsedDocument, engagement: dict[str, Any]) -> list[dict[str, Any]]:
    notes: list[dict[str, Any]] = []
    text = doc.text
    low = text.lower()
    sheet_no, sheet_title = re_engine._doc_sheet(doc)
    doc_account = re_engine._account_from_code(sheet_no) or re_engine._sanitize_account_label(sheet_title)

    prep = engagement.get("preparer", "")
    rev = engagement.get("reviewer", "")
    if prep and rev and prep not in ("확인 필요",) and prep == rev:
        notes.append(_qc_note(
            defect="작성자와 검토자(심리자)가 동일인",
            reason="이중검토 원칙에 부합하지 않을 수 있습니다.",
            to_be="검토자를 작성자와 다른 회계사로 지정하십시오.",
            sheet_no=sheet_no, sheet_title=sheet_title, importance="중",
        ))

    notes += _going_concern_notes(doc)
    notes += _going_concern_impairment_notes(doc)
    notes = prune_documented_review_notes(doc, notes)

    if any(k in low for k in _INDEPENDENCE_KW):
        if not re.search(r"독립성\s*확인|위협\s*평가", low):
            notes.append(_qc_note(
                defect="독립성 관련 언급 — 위협 평가 미확인",
                reason="독립성·비감사 관련 서술이 있으나 대응조치 문서화가 보이지 않습니다.",
                to_be="독립성 위협·대응조치를 문서화하십시오.",
                sheet_no=sheet_no, sheet_title=doc_account or sheet_title, importance="중",
            ))

    notes += _check_cross_account_consistency(doc)
    return notes


def _excel_ref(table, row_i: int, col_i: int) -> str:
    row_map = table.attrs.get("row_map") or []
    col_map = table.attrs.get("col_map") or []
    row_no = int(row_map[row_i]) if row_i < len(row_map) else row_i + 1
    col_no = int(col_map[col_i]) if col_i < len(col_map) else col_i + 1
    return f"{get_column_letter(col_no)}{row_no}"


def _classify_gc_scope(
    text: str,
    sheet_no: str,
    sheet_title: str,
    account: str,
) -> str:
    ctx = f"{sheet_no} {sheet_title} {account} {text}"
    scores = {label: 0 for label in _SCOPE_LABELS if label != "미확인(범위 명시 필요)"}
    if _SCOPE_BORROWER_RE.search(ctx):
        scores["대여(차입)처·피대여 회사"] += 2
    if _SCOPE_SUB_RE.search(ctx):
        scores["종속·관계회사"] += 2
    if _SCOPE_AUDIT_RE.search(ctx):
        scores["감사대상 회사"] += 2
    if _LOAN_SHEET_CODES.search(sheet_no or ""):
        scores["대여(차입)처·피대여 회사"] += 2
        if "손상" in sheet_title or "대여" in f"{account}{sheet_title}":
            scores["대여(차입)처·피대여 회사"] += 1
    if _SUB_SHEET_CODES.search(sheet_no or ""):
        scores["종속·관계회사"] += 2
    best_score = max(scores.values())
    if best_score == 0:
        return "미확인(범위 명시 필요)"
    top = [k for k, v in scores.items() if v == best_score]
    if len(top) > 1:
        return "미확인(범위 명시 필요)"
    return top[0]


def _scope_explicitly_documented(full_text: str, scope: str) -> bool:
    if "미확인" in scope:
        return False
    patterns = {
        "감사대상 회사": r"피감사|감사대상|보고기업|연결실체|본회사",
        "종속·관계회사": r"종속|관계회사|관계기업|지분법",
        "대여(차입)처·피대여 회사": r"피대여|대여처|차입처|채무자|차주",
    }
    p = patterns.get(scope)
    if not p:
        return False
    return bool(re.search(p, full_text, re.I) and _GC_ASSUMPTION_RE.search(full_text))


def _scan_gc_mentions(doc: ParsedDocument) -> list[_GCMention]:
    mentions: list[_GCMention] = []
    seen: set[tuple[str, int, int, str]] = set()
    for t in doc.tables:
        sheet_no = str(t.attrs.get("source", "")).strip()
        sheet_title = str(t.attrs.get("title", "")).strip()
        account = re_engine.table_account(t) or re_engine._sanitize_account_label(sheet_title)
        for ri, row in enumerate(t.itertuples(index=False)):
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                if isinstance(cell, float) and cell != cell:
                    continue
                txt = str(cell).strip()
                if len(txt) < 4 or not _GC_NARRATIVE_RE.search(txt):
                    continue
                key = (sheet_no, ri, ci, txt[:40])
                if key in seen:
                    continue
                seen.add(key)
                scope = _classify_gc_scope(txt, sheet_no, sheet_title, account or "")
                mentions.append(_GCMention(
                    sheet_no=sheet_no or "-",
                    sheet_title=sheet_title or account or "-",
                    excel_ref=_excel_ref(t, ri, ci),
                    excerpt=txt[:120],
                    scope=scope,
                ))
    if not mentions and _GC_NARRATIVE_RE.search(doc.text):
        sheet_no, sheet_title = re_engine._doc_sheet(doc)
        scope = _classify_gc_scope(doc.text[:3000], sheet_no, sheet_title, "")
        match = _GC_NARRATIVE_RE.search(doc.text)
        mentions.append(_GCMention(
            sheet_no=sheet_no or "-",
            sheet_title=sheet_title or "-",
            excel_ref="(본문)",
            excerpt=match.group(0) if match else "",
            scope=scope,
        ))
    return mentions


def _going_concern_notes(doc: ParsedDocument) -> list[dict[str, Any]]:
    """계속기업 관련 서술이 있을 때 범위·위치·미비 조서를 구체적으로 지적."""
    mentions = _filter_undocumented_mentions(doc, _scan_gc_mentions(doc))
    if not mentions:
        return []

    full = doc.text
    has_assumption_review = bool(_GC_ASSUMPTION_RE.search(full))

    by_scope: dict[str, list[_GCMention]] = {}
    for m in mentions:
        by_scope.setdefault(m.scope, []).append(m)

    notes: list[dict[str, Any]] = []
    for scope, items in by_scope.items():
        loc_parts = [f"{_short_sheet(m.sheet_no)} {m.excel_ref}" for m in items[:5]]
        location = " · ".join(loc_parts)
        if len(items) > 5:
            location += f" 외 {len(items) - 5}건"

        gaps: list[str] = []
        if "미확인" in scope:
            gaps.append(
                "계속기업 관련 서술은 있으나, 검토 대상이 감사대상 회사·종속·관계회사·"
                "대여(차입)처 중 어디인지 조서에 명시되어 있지 않습니다."
            )
        elif not _scope_explicitly_documented(full, scope):
            gaps.append(
                f"계속기업 관련 서술은 「{scope}」에 대한 것으로 보이나, "
                f"해당 범위의 계속기업 가정 검토가 명시·문서화되어 있지 않습니다."
            )

        if not has_assumption_review:
            gaps.append(
                "KSA 570에 따른 계속기업 가정 적정성 검토·"
                "중요한 의문(불확실성) 평가 및 결론이 확인되지 않습니다."
            )
        elif not _scope_explicitly_documented(full, scope) and "미확인" not in scope:
            gaps.append(
                f"문서에 계속기업 가정 검토 서술은 있으나, "
                f"본 서술이 대상하는 「{scope}」 범위와의 대응 관계가 불명확합니다."
            )

        if not gaps:
            continue

        ref_items = [m for m in items if m.excel_ref != "(본문)"]
        if (
            items
            and ref_items
            and all(
                _location_is_documented(doc, _short_sheet(m.sheet_no), m.excel_ref)
                for m in ref_items
            )
            and _sheet_has_substantive_review(doc, _short_sheet(items[0].sheet_no))
        ):
            continue

        if not items:
            continue

        note_sheet_no, note_sheet_title = _first_sheet_meta(items)

        loc_detail = "; ".join(
            f"「{m.sheet_no}」 {m.excel_ref} — {m.excerpt[:60]}"
            for m in items[:3]
        )
        scope_label = (
            "감사대상 회사 / 종속·관계회사 / 대여(차입)처 중 해당 범위"
            if "미확인" in scope
            else scope
        )

        notes.append(_qc_note(
            defect=f"계속기업 가정 검토 — {scope_label} 범위·문서화 보완 필요",
            reason=(
                f"조서에 계속기업·유동성·자본잠식 등 관련 서술이 있습니다 ({location}). "
                + " ".join(gaps)
                + f" 관련 위치: {loc_detail}."
            ),
            to_be=(
                f"① 조서번호·행열({location})에 해당하는 계속기업 서술의 검토 대상을 "
                f"「{scope_label}」로 명확히 정의하십시오. "
                "② 해당 범위에 대해 계속기업 가정 적정성·중요한 의문 여부·"
                "대손/손상·공시 영향 등 관련 조서의 검토 절차와 결론을 구체적으로 문서화하십시오."
            ),
            sheet_no=note_sheet_no,
            sheet_title=note_sheet_title,
            importance="상",
            location=location,
            basis="회계감사기준 570호(계속기업); 품질관리기준",
        ))
    return notes


def _has_gc_uncertainty(doc: ParsedDocument, mentions: list[_GCMention]) -> bool:
    if _GC_UNCERTAINTY_RE.search(doc.text):
        return True
    return any(_GC_UNCERTAINTY_RE.search(m.excerpt) for m in mentions)


def _doc_accounts(doc: ParsedDocument) -> set[str]:
    accts: set[str] = set()
    for t in doc.tables:
        acct = re_engine.table_account(t) or re_engine._sanitize_account_label(
            str(t.attrs.get("title", ""))
        )
        if acct:
            accts.add(acct)
    return accts


def _doc_text_hits(doc: ParsedDocument, patterns: tuple[str, ...]) -> bool:
    full = doc.text
    return any(re.search(p, full, re.I) for p in patterns)


def _short_sheet(sheet_no: str) -> str:
    return out_fmt.short_sheet_code(sheet_no) or sheet_no


def _allowance_omission_justified(
    cell: str, row: str = "", block: str = ""
) -> bool:
    """추가 대손 미반영 결론이 인접 맥락에서 회수가능성·중요성 등으로 소명된 경우."""
    cell_t = str(cell or "").strip()
    row_t = str(row or "").strip()
    block_t = str(block or "").strip()
    combined = " ".join(x for x in (cell_t, row_t, block_t) if x)
    if not _NO_ADDITIONAL_ALLOWANCE_CONCLUSION_RE.search(combined):
        return False
    rationale = " ".join(x for x in (row_t, block_t, cell_t) if x)
    return bool(_ALLOWANCE_OMISSION_JUSTIFICATION_RE.search(rationale))


def _context_has_substantive_gap(
    cell: str, row: str = "", block: str = ""
) -> bool:
    """셀·인접 맥락 기준 중대한 절차누락·명시적 미비 여부."""
    if _allowance_omission_justified(cell, row, block):
        return False
    return _cell_has_substantive_gap(cell)


def _has_documented_review(cell_text: str, row_text: str = "", *, block_text: str = "") -> bool:
    """셀·인접 행·블록에 검토·판단·결론 서술이 있으면 지적 대상에서 제외."""
    txt = str(cell_text or "").strip()
    if _allowance_omission_justified(txt, row_text, block_text):
        return True
    if txt and _context_has_substantive_gap(txt, row_text, block_text):
        return False
    for hay in (txt, row_text, block_text):
        if not hay or len(hay.strip()) < 8:
            continue
        if _EXPLICIT_GAP_RE.search(hay) and hay == txt:
            continue
        if _REVIEW_DOCUMENTED_RE.search(hay):
            return True
    return False


def _sheet_text_by_code(doc: ParsedDocument, sheet_code: str) -> str:
    key = sheet_code.upper()
    cache = _doc_cache(doc)["sheet_text"]
    if key in cache:
        return cache[key]
    code_up = key
    parts: list[str] = []
    for t in doc.tables:
        sn = str(t.attrs.get("source", ""))
        if out_fmt.short_sheet_code(sn).upper() == code_up or code_up in sn.upper():
            parts.append(re_engine._sheet_text(t))
    text = "\n".join(parts)
    cache[key] = text
    return text


def _sheet_has_substantive_review(doc: ParsedDocument, sheet_code: str) -> bool:
    """시트에 감사메모·절차 틀 + 다수의 검토·판단 서술이 있으면 문서화된 것으로 본다."""
    key = sheet_code.upper()
    cache = _doc_cache(doc)["subst_review"]
    if key in cache:
        return cache[key]
    text = _sheet_text_by_code(doc, sheet_code)
    if len(text.strip()) < 120:
        cache[key] = False
        return False
    if not _SHEET_REVIEW_FRAME_RE.search(text):
        hits = len(_REVIEW_DOCUMENTED_RE.findall(text))
        ok = hits >= 6
    else:
        ok = len(_REVIEW_DOCUMENTED_RE.findall(text)) >= 4
    cache[key] = ok
    return ok


def _block_text_near_ref(
    doc: ParsedDocument, sheet_code: str, excel_ref: str, *, window: int = 8
) -> str:
    m = re.match(r"([A-Z]+)(\d+)", excel_ref.strip(), re.I)
    if not m:
        return ""
    cache = _doc_cache(doc)["blocks"]
    cache_key = (sheet_code.upper(), excel_ref.upper(), window)
    if cache_key in cache:
        return cache[cache_key]
    target_row = int(m.group(2))
    rows = _sheet_rows(doc, sheet_code)
    chunks = [txt for row_no, txt in rows.items() if abs(row_no - target_row) <= window]
    result = " ".join(chunks)
    cache[cache_key] = result
    return result


def _location_is_documented(doc: ParsedDocument, sheet_code: str, excel_ref: str) -> bool:
    """지적 위치(조서·셀)에 검토내역이 있거나 인접 블록·시트 전반이 문서화된 경우."""
    if excel_ref in ("(본문)", ""):
        return _sheet_has_substantive_review(doc, sheet_code)
    cell, row = _cell_at_ref(doc, sheet_code, excel_ref)
    block = _block_text_near_ref(doc, sheet_code, excel_ref)
    if _allowance_omission_justified(cell, row, block):
        return True
    if cell and _context_has_substantive_gap(cell, row, block):
        return False
    if _has_documented_review(cell, row, block_text=block):
        return True
    if _sheet_has_substantive_review(doc, sheet_code) and block and len(block) > 60:
        if _REVIEW_DOCUMENTED_RE.search(block):
            return True
    return False


def _filter_undocumented_mentions(
    doc: ParsedDocument, mentions: list[_GCMention]
) -> list[_GCMention]:
    kept: list[_GCMention] = []
    for m in mentions:
        short = _short_sheet(m.sheet_no)
        if m.excel_ref != "(본문)" and _location_is_documented(doc, short, m.excel_ref):
            continue
        kept.append(m)
    return kept


def _cell_has_substantive_gap(cell: str) -> bool:
    """셀에 중대한 절차누락·명시적 미비가 있는지 (회계처리 결론의 '반영하지 않음'은 제외)."""
    txt = str(cell or "").strip()
    if not txt:
        return False
    if _REVIEW_DOCUMENTED_RE.search(txt) and re.search(
        r"분류함|판단함|검토함|확인함", txt, re.I
    ):
        if not re.search(r"추가\s*대손|대손충당|손상.*반영|미반영", txt, re.I):
            return False
    return bool(_EXPLICIT_GAP_RE.search(txt))


_SUBSTANTIVE_GAP_IN_NOTE_RE = re.compile(
    r"미반영|반영.*않|누락|미수행|부족|과소|불일치|중대|절차\s*누락|흔적\s*미확인",
    re.I,
)
_TOPIC_NOTE_CHECKS: list[tuple[re.Pattern[str], re.Pattern[str]]] = [
    (
        re.compile(r"지분법", re.I),
        re.compile(
            r"지분법.{0,120}(?:검토|판단|분류|손상)|"
            r"(?:검토|판단|분류).{0,80}지분법",
            re.I,
        ),
    ),
    (
        re.compile(r"장기대여금|대여금|대여\s*채권", re.I),
        re.compile(
            r"(?:손상|회수|대손).{0,80}(?:검토|판단|산정)|손상\s*검토",
            re.I,
        ),
    ),
    (
        re.compile(r"회수\s*가능|회수가능", re.I),
        re.compile(r"회수\s*가능|회수가능|사용가치|공정가치|회수가능액", re.I),
    ),
    (
        re.compile(r"손상기준|손상\s*검토", re.I),
        re.compile(r"손상기준|손상\s*징후|손상\s*검토", re.I),
    ),
    (
        re.compile(r"평가|공시", re.I),
        re.compile(
            r"(?:평가|공정가치|공시).{0,100}(?:검토|판단|적정|산정)|"
            r"감사(?:메모|절차)",
            re.I,
        ),
    ),
]
_ACCOUNT_EXTRA_SYNS: dict[str, tuple[str, ...]] = {
    "투자자산": ("장기대여금", "대여금", "손상검토", "투자부동산"),
}


def _account_sheet_synonyms(account: str) -> list[str]:
    syns = [account]
    for name, items in re_engine.ACCOUNT_TAXONOMY:
        if name == account:
            syns.extend(items)
    syns.extend(_ACCOUNT_EXTRA_SYNS.get(account, ()))
    return list(dict.fromkeys(syns))


def _account_sheets_text(doc: ParsedDocument, account: str) -> str:
    """계정과목에 해당하는 조서 본문을 통합."""
    syns = _account_sheet_synonyms(account)
    parts: list[str] = []
    seen: set[str] = set()
    for t in doc.tables:
        sn = str(t.attrs.get("source", "")).strip()
        if sn in seen:
            continue
        title = str(t.attrs.get("title", "")).strip()
        sheet_text = re_engine._sheet_text(t)
        acct = re_engine.table_account(t)
        matched = acct == account
        if not matched:
            ctx = f"{sn} {title}"
            matched = any(s in ctx or s in sheet_text[:1000] for s in syns)
        if not matched and account == "투자자산":
            matched = bool(re.search(r"/F\d{3}|SAJ", sn, re.I))
        if matched:
            parts.append(sheet_text)
            seen.add(sn)
    return "\n".join(parts)


def _text_has_substantive_review(text: str) -> bool:
    """통합 본문에 감사 절차·검토 서술이 충분한지."""
    if len(text.strip()) < 120:
        return False
    hits = len(_REVIEW_DOCUMENTED_RE.findall(text))
    if _SHEET_REVIEW_FRAME_RE.search(text):
        return hits >= 4
    return hits >= 6


def _note_topic_covered(note_hay: str, scope_text: str) -> bool:
    """지적 주제가 조서 본문에 이미 검토·문서화되어 있는지."""
    matched = [(n, e) for n, e in _TOPIC_NOTE_CHECKS if n.search(note_hay)]
    if not matched:
        return False
    return all(e.search(scope_text) for _, e in matched)


def _brief_bullets_documented(doc: ParsedDocument, note: dict[str, Any]) -> bool:
    """통합 검토요청 노트의 보완지침 항목이 모두 조서에서 해소되었는지."""
    to_be = str(note.get("to_be") or "")
    bullets = [
        re.sub(r"^[·\-\s]+", "", ln).strip()
        for ln in to_be.splitlines()
        if re.match(r"^\s*[·\-]", ln)
    ]
    if not bullets:
        return False
    acct = re_engine.note_account(note) or ""
    scope = _account_sheets_text(doc, acct) if acct else (doc.text or "")
    if not _text_has_substantive_review(scope):
        return False
    return all(_note_topic_covered(b, scope) for b in bullets)


def _note_should_suppress(doc: ParsedDocument, note: dict[str, Any]) -> bool:
    """검토내역이 충분히 문서화된 지적은 제외 (명시적 중대 미비만 유지)."""
    hay = f"{note.get('defect', '')} {note.get('reason', '')} {note.get('to_be', '')}"

    if note.get("brief_merged") and _brief_bullets_documented(doc, note):
        return True

    acct = re_engine.note_account(note)
    if acct:
        scope = _account_sheets_text(doc, acct)
        if scope and _text_has_substantive_review(scope) and _note_topic_covered(hay, scope):
            return True

    refs = out_fmt.parse_location_refs(str(note.get("location") or ""))
    if refs:
        return all(_location_is_documented(doc, sheet, ref) for sheet, ref in refs)

    sheet = _short_sheet(str(note.get("sheet_no") or note.get("workpaper_ref") or ""))
    if sheet:
        sheet_text = _sheet_text_by_code(doc, sheet)
        if sheet_text and _text_has_substantive_review(sheet_text) and _note_topic_covered(hay, sheet_text):
            return True
    if sheet and _sheet_has_substantive_review(doc, sheet):
        if note.get("category") in ("QC·품질관리", "검토요청", "증빙·절차", "개선제안"):
            return True
        if note.get("source") == "ai" and _REVIEW_DOCUMENTED_RE.search(
            _sheet_text_by_code(doc, sheet)
        ):
            return True

    if note.get("enforcement_protected") or (
        note.get("category") == "감리지적체크" and note.get("enforcement_checklist_id")
    ):
        return _enforcement_note_documented(doc, note)

    if note.get("focus_protected") or note.get("category") == "중점감리":
        return _focus_note_documented(doc, note)

    if _EXPLICIT_GAP_RE.search(hay) and _SUBSTANTIVE_GAP_IN_NOTE_RE.search(hay):
        return False
    if note.get("category") == "절차누락" and note.get("importance") == "상":
        if re.search(r"흔적\s*미확인|미수행|누락", hay) and not _REVIEW_DOCUMENTED_RE.search(hay):
            return False

    return False


def _focus_note_documented(doc: ParsedDocument, note: dict[str, Any]) -> bool:
    """4대 중점 노트 — 해당 checklist_id 항목이 조서에서 충족되면 제외."""
    try:
        import fss_focus
        import guidelines_loader as gl
    except ImportError:
        return False
    issue_no = note.get("focus_issue_no")
    checklist_id = str(note.get("focus_checklist_id") or "").strip()
    if not issue_no:
        return False
    listed = note.get("is_listed")
    if listed is None:
        cid = str(note.get("focus_checklist_id") or "")
        if cid.startswith("U"):
            listed = False
        elif cid.startswith("L"):
            listed = True
        else:
            listed = True
    else:
        listed = bool(listed)

    gl.load_focus_issues_from_db.cache_clear()
    enhanced = gl.load_focus_issues_from_db(listed)
    items_by_issue: dict[int, list] = {}
    issues_meta: dict[int, Any] = {}
    if enhanced:
        for e in enhanced:
            items_by_issue[e.issue_no] = list(e.checklist)
            issues_meta[e.issue_no] = e
    else:
        for issue in fss_focus.load_current_focus_issues(2026, listed):
            items_by_issue[issue.issue_no] = list(issue.checklist)
            issues_meta[issue.issue_no] = issue

    target_items = items_by_issue.get(issue_no, [])
    if checklist_id:
        target_items = [
            it for it in target_items
            if str(getattr(it, "checklist_id", "") or "").strip() == checklist_id
        ]
    if not target_items:
        return False

    issue = issues_meta.get(issue_no)
    if issue is None:
        return False

    for item in target_items:
        matched = fss_focus._matched_sheets_for_item(doc, issue, item)
        if not matched:
            return False
        body = "\n".join(s[2] for s in matched)
        if not fss_focus._focus_checklist_satisfied(body, issue, item):
            return False
    return bool(target_items)


def _enforcement_note_documented(doc: ParsedDocument, note: dict[str, Any]) -> bool:
    """감리지적 체크리스트 노트 — 동일 checklist_id 항목이 조서에서 충족되면 제외."""
    try:
        import enforcement_review as er
        import guidelines_loader as gl
    except ImportError:
        return False

    checklist_id = str(note.get("enforcement_checklist_id") or "").strip()
    if not checklist_id:
        return False

    sheet = out_fmt.short_sheet_code(str(note.get("sheet_no") or note.get("workpaper_ref") or "")).upper()
    if not sheet:
        return False

    listed = note.get("is_listed")
    if listed is None:
        listed = True
    else:
        listed = bool(listed)

    items = gl.load_enforcement_checklist_from_db(listed) or []
    target = [it for it in items if str(it.checklist_id or "").strip() == checklist_id]
    if not target:
        return False
    item = target[0]

    tables = er._tables_for_code(doc, sheet)
    if not tables:
        return False
    stext = "\n".join(re_engine._sheet_text(t) for t in tables)
    return gl.checklist_satisfied_enhanced(stext, item)


def _row_text(row) -> str:
    return " ".join(str(c) for c in row if c is not None and str(c).strip())


def _doc_cache(doc: ParsedDocument) -> dict[str, Any]:
    cache = getattr(doc, "_qc_cache", None)
    if cache is None:
        cache = {
            "sheet_rows": {},
            "sheet_text": {},
            "subst_review": {},
            "blocks": {},
            "cells": {},
        }
        setattr(doc, "_qc_cache", cache)
    return cache


def _sheet_rows(doc: ParsedDocument, sheet_code: str) -> dict[int, str]:
    """시트 코드별 행번호→행 텍스트 (1회 구축·캐시)."""
    key = sheet_code.upper()
    cache = _doc_cache(doc)["sheet_rows"]
    if key in cache:
        return cache[key]
    rows: dict[int, str] = {}
    for t in doc.tables:
        sn = str(t.attrs.get("source", ""))
        short = out_fmt.short_sheet_code(sn).upper()
        if short != key and key not in sn.upper():
            continue
        row_map = t.attrs.get("row_map") or []
        for ri, row in enumerate(t.itertuples(index=False, name=None)):
            row_no = int(row_map[ri]) if ri < len(row_map) else ri + 1
            rows[row_no] = _row_text(row)
    cache[key] = rows
    return rows


def _cell_has_flaw_signal(txt: str) -> bool:
    return bool(
        _INSUFFICIENT_ALLOWANCE_RE.search(txt)
        or _LOGICAL_FLAW_RE.search(txt)
        or any(rule.search(txt) for rule, _ in _FLAW_CELL_RULES)
    )


def _scan_impairment_hits(doc: ParsedDocument) -> list[_ImpairmentHit]:
    hits: list[_ImpairmentHit] = []
    seen: set[tuple[str, str, str]] = set()
    doc_gc = bool(_GC_UNCERTAINTY_RE.search(doc.text))
    for t in doc.tables:
        sheet_no = str(t.attrs.get("source", "")).strip()
        sheet_title = str(t.attrs.get("title", "")).strip()
        account = re_engine.table_account(t) or re_engine._sanitize_account_label(sheet_title)
        ctx = f"{sheet_no} {sheet_title} {account}"
        short_code = out_fmt.short_sheet_code(sheet_no) or sheet_no
        if _LOAN_SHEET_CODES.search(sheet_no):
            default_cat = "loan"
        elif _SUB_SHEET_CODES.search(sheet_no) or re.search(r"지분법|투자주식|관계기업", ctx, re.I):
            default_cat = "equity"
        elif re.search(r"매출채권|대손|외상매출|받을어음", ctx, re.I):
            default_cat = "allowance"
        else:
            default_cat = "allowance"

        sheet_txt_head = re_engine._sheet_text(t)[:2000]
        sheet_has_gc = bool(_GC_UNCERTAINTY_RE.search(sheet_txt_head))
        row_map = t.attrs.get("row_map") or []

        for ri, row in enumerate(t.itertuples(index=False, name=None)):
            row_txt = _row_text(row)
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                if isinstance(cell, float) and cell != cell:
                    continue
                txt = str(cell).strip()
                if len(txt) < 6 or not _cell_has_flaw_signal(txt):
                    continue
                if not (_GC_UNCERTAINTY_RE.search(txt) or _GC_NARRATIVE_RE.search(txt)):
                    if not (
                        _GC_UNCERTAINTY_RE.search(row_txt)
                        or sheet_has_gc
                        or doc_gc
                    ):
                        continue

                ref = _excel_ref(t, ri, ci)
                block = _block_text_near_ref(doc, short_code, ref)
                if _has_documented_review(txt, row_txt, block_text=block):
                    continue
                if (
                    _sheet_has_substantive_review(doc, short_code)
                    and block
                    and len(block) > 60
                    and _REVIEW_DOCUMENTED_RE.search(block)
                    and not _context_has_substantive_gap(txt, row_txt, block)
                ):
                    continue

                issue = ""
                for rule, label in _FLAW_CELL_RULES:
                    if rule.search(txt):
                        issue = label
                        break
                if not issue:
                    if _INSUFFICIENT_ALLOWANCE_RE.search(txt):
                        issue = "대손충당금·손상반영이 부족하거나 미반영으로 보임"
                    else:
                        issue = "충당금·손상 산정에 논리적 결함이 있어 보임"

                key = (sheet_no, ref, issue)
                if key in seen:
                    continue
                seen.add(key)
                cat = default_cat
                if re.search(r"지분법|투자주식|관계기업|종속", txt, re.I):
                    cat = "equity"
                elif re.search(r"장기대여|대여금|대손충당금\(장기", txt, re.I):
                    cat = "loan"
                hits.append(_ImpairmentHit(
                    sheet_no=sheet_no or "-",
                    sheet_title=sheet_title or account or "-",
                    excel_ref=ref,
                    excerpt=txt[:120],
                    issue=issue,
                    category=cat,
                ))
    return hits


def _missing_impairment_reviews(
    doc: ParsedDocument,
    accts: set[str],
) -> list[tuple[str, str]]:
    """계속기업 불확실성 대비 누락된 연계 검토 (category, message)."""
    missing: list[tuple[str, str]] = []
    full = doc.text

    has_receivable = (
        bool(accts & {"매출채권", "대손충당금"})
        or _doc_text_hits(doc, (r"매출채권", r"외상매출금", r"받을어음"))
    )
    if has_receivable and not _ALLOWANCE_REVIEW_RE.search(full):
        missing.append((
            "allowance",
            "계속기업 불확실성이 있으나 매출채권·대손충당금(기대신용손실) "
            "적정성 검토 절차·산정 근거가 확인되지 않습니다.",
        ))

    has_equity = (
        bool(accts & {"지분법", "투자주식", "종속기업", "관계기업"})
        or _SUB_SHEET_CODES.search(full)
        or _doc_text_hits(doc, (r"지분법", r"투자주식", r"관계기업", r"종속기업"))
    )
    if has_equity and not _EQUITY_IMPAIR_RE.search(full):
        missing.append((
            "equity",
            "계속기업 불확실성이 있으나 관련 투자주식·지분법 적용 주식의 "
            "손상평가·회수가능액 검토가 확인되지 않습니다.",
        ))

    has_loan = (
        bool(accts & {"장기대여금", "대여금"})
        or any(_LOAN_SHEET_CODES.search(str(t.attrs.get("source", ""))) for t in doc.tables)
        or _doc_text_hits(doc, (r"장기대여금", r"대여금\s*손상"))
    )
    if has_loan and not re.search(r"손상|회수가능|대손|충당", full, re.I):
        missing.append((
            "loan",
            "피대여 회사의 계속기업 불확실성이 있으나 장기대여금·대여채권 "
            "손상검토·대손충당금 설정 근거가 확인되지 않습니다.",
        ))
    return missing


def _going_concern_impairment_notes(doc: ParsedDocument) -> list[dict[str, Any]]:
    """계속기업 불확실성 시 대손충당금·지분법 손상 등 연계 검토 및 논리 결함 지적."""
    mentions = _scan_gc_mentions(doc)
    if not mentions and not _GC_UNCERTAINTY_RE.search(doc.text):
        return []
    if not _has_gc_uncertainty(doc, mentions):
        return []

    accts = _doc_accounts(doc)
    hits = _scan_impairment_hits(doc)
    missing = _missing_impairment_reviews(doc, accts)

    by_category: dict[str, list[_ImpairmentHit]] = {}
    for h in hits:
        by_category.setdefault(h.category, []).append(h)

    notes: list[dict[str, Any]] = []
    category_labels = {
        "allowance": "대손충당금(기대신용손실)",
        "equity": "지분법·투자주식 손상",
        "loan": "장기대여금·대여채권 손상",
    }
    category_accounts = {
        "allowance": "대손충당금",
        "equity": "지분법",
        "loan": "장기대여금",
    }

    for cat, label in category_labels.items():
        cat_hits = by_category.get(cat, [])
        cat_missing = [msg for c, msg in missing if c == cat]
        if not cat_hits and not cat_missing:
            continue

        loc_parts = [
            f"{_short_sheet(h.sheet_no)} {h.excel_ref}" for h in cat_hits[:5]
        ]
        location = " · ".join(loc_parts) if loc_parts else "-"
        if len(cat_hits) > 5:
            location += f" 외 {len(cat_hits) - 5}건"

        reasons: list[str] = []
        if cat_missing:
            reasons.extend(cat_missing)
        for h in cat_hits[:3]:
            short = _short_sheet(h.sheet_no)
            reasons.append(f"「{short}」 {h.excel_ref}: {h.issue} — {h.excerpt[:50]}")

        if not reasons:
            continue

        sheet_no, sheet_title = _first_sheet_meta(
            cat_hits,
            mentions,
            default_title=category_accounts.get(cat) or label,
        )

        notes.append(_qc_note(
            defect=f"계속기업 불확실성 — {label} 검토·조서 보완 필요",
            reason="검토대상 회사에 계속기업 불확실성이 확인됩니다. " + " ".join(reasons),
            to_be=(
                f"① {label}이 계속기업 불확실성을 충분히 반영하도록 "
                "산정 근거·가정·민감도 분석을 조서에 보완하십시오. "
                f"② 조서({location})의 논리적 결함·충당금 부족이 의심되는 부분을 "
                "재검토하고 필요 시 추가 대손충당금·손상차손 반영 여부를 문서화하십시오."
            ),
            sheet_no=sheet_no,
            sheet_title=sheet_title,
            importance="상",
            location=location,
            basis="회계감사기준 570호(계속기업); K-IFRS 제1109호·제1036호(손상)",
        ))
    return notes


def _cell_at_ref(doc: ParsedDocument, sheet_code: str, excel_ref: str) -> tuple[str, str]:
    """조서 코드·셀좌표에 해당하는 셀 본문·행 맥락."""
    m = re.match(r"([A-Z]+)(\d+)", excel_ref.strip(), re.I)
    if not m:
        return "", ""
    cache = _doc_cache(doc)["cells"]
    cache_key = (sheet_code.upper(), excel_ref.upper())
    if cache_key in cache:
        return cache[cache_key]
    target_col = m.group(1).upper()
    target_row = int(m.group(2))
    code_up = sheet_code.upper()

    for t in doc.tables:
        sn = str(t.attrs.get("source", ""))
        short = out_fmt.short_sheet_code(sn)
        if code_up not in sn.upper() and short.upper() != code_up:
            continue
        row_map = t.attrs.get("row_map") or []
        col_map = t.attrs.get("col_map") or []
        for ri, row in enumerate(t.itertuples(index=False, name=None)):
            row_no = int(row_map[ri]) if ri < len(row_map) else ri + 1
            if row_no != target_row:
                continue
            for ci, cell in enumerate(row):
                col_no = int(col_map[ci]) if ci < len(col_map) else ci + 1
                if get_column_letter(col_no) != target_col:
                    continue
                cell_txt = "" if cell is None else str(cell).strip()
                result = (cell_txt, _row_text(row))
                cache[cache_key] = result
                return result
    result = ("", "")
    cache[cache_key] = result
    return result


def prune_documented_review_notes(
    doc: ParsedDocument, notes: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """검토내역이 있는 위치·시트는 지적 제외 — 중대한 절차누락·명시적 미비만 유지."""
    return [n for n in notes if not _note_should_suppress(doc, n)]


def _check_cross_account_consistency(doc: ParsedDocument) -> list[dict[str, Any]]:
    notes: list[dict[str, Any]] = []
    pairs = [
        ("매출채권", "매출", "매출채권 조서는 있으나 매출 관련 서술이 확인되지 않습니다."),
        ("차입금", "이자", "차입금 조서는 있으나 이자비용 관련 서술이 확인되지 않습니다."),
    ]
    accts: set[str] = set()
    for t in doc.tables:
        title = str(t.attrs.get("title", ""))
        acct = re_engine.table_account(t) or title
        if acct:
            accts.add(acct)
    full = doc.text
    for acct_a, kw_b, msg in pairs:
        if acct_a not in accts or re.search(kw_b, full, re.IGNORECASE):
            continue
        notes.append(_qc_note(
            defect=f"조서 상호정합 — {acct_a} vs {kw_b}",
            reason=msg,
            to_be=f"{acct_a}와 {kw_b} 관련 조서 연결·대사를 확인하십시오.",
            sheet_no="-", sheet_title=acct_a, importance="중",
        ))
    return notes


def _qc_note(
    *,
    defect: str,
    reason: str,
    to_be: str,
    sheet_no: str,
    sheet_title: str,
    importance: str = "중",
    location: str = "",
    basis: str = "품질관리기준; 회계감사기준 220호",
) -> dict[str, Any]:
    return {
        "id": "",
        "importance": importance,
        "category": "QC·품질관리",
        "defect": defect,
        "reason": reason,
        "basis": basis,
        "to_be": to_be,
        "sheet_no": sheet_no or "-",
        "sheet_title": sheet_title,
        "sheet": f"{sheet_no} ({sheet_title})" if sheet_no and sheet_title else sheet_no or sheet_title,
        "location": location,
        "summary": "",
        "workpaper_ref": sheet_no or "-",
        "source": "rule",
    }
