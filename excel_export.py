"""감리 리뷰노트 엑셀 생성.

SKILL_감리리뷰노트_자동생성 스킬 양식을 따른다.
  · 표지 · 우선순위 목록 · 등급별(★★★/★★/★) 세부 시트
규칙엔진이 만든 리뷰노트를 (등급/조서/제목/상세/기준/보완지침) 구조로 매핑한다.
"""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from typing import Any

import output_formatter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 색상 팔레트 (감리 리뷰노트 표준) ──────────────────────────────
COLOR = {
    "header_bg": "1F3864",
    "header_font": "FFFFFF",
    "tier1_bg": "C00000",
    "tier1_font": "FFFFFF",
    "tier2_bg": "F4B942",
    "tier2_font": "3B2B00",
    "tier3_bg": "FFE699",
    "tier3_font": "7F5000",
    "guide_bg": "EBF7EE",
    "label_bg": "EBF3FB",
    "divider": "D9D9D9",
}

_FONT = "맑은 고딕"


# ── 스타일 헬퍼 ──────────────────────────────────────────────
def _bdr(style: str = "thin") -> Border:
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _fl(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def _fn(bold: bool = False, sz: int = 10, color: str = "000000", italic: bool = False) -> Font:
    return Font(name=_FONT, bold=bold, size=sz, color=color, italic=italic)


def _al(h: str = "left", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _mc(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _set(ws, row, col, val, *, bold=False, sz=10, color="000000",
         fill=None, halign="left", wrap=True, italic=False, border=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _fn(bold=bold, sz=sz, color=color, italic=italic)
    c.alignment = _al(halign, "center", wrap)
    if fill:
        c.fill = _fl(fill)
    if border:
        c.border = _bdr()
    return c


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _no_grid(ws) -> None:
    ws.sheet_view.showGridLines = False


def _divider(ws, row: int, num_cols: int = 6, start_col: int = 2) -> None:
    ws.row_dimensions[row].height = 6
    for c in range(start_col, start_col + num_cols):
        ws.cell(row=row, column=c).fill = _fl(COLOR["divider"])


TIER_STAR = {1: "★★★\n즉시", 2: "★★\n조기", 3: "★\n단기"}
TIER_LABEL = {1: "★★★ 즉시", 2: "★★ 조기", 3: "★ 단기"}
TIER_NAME = {1: "★★★ 즉시 수정", 2: "★★ 조기 보완", 3: "★ 단기 보완"}
TIER_TITLE = {1: "★★★ 즉시수정", 2: "★★ 조기보완", 3: "★ 단기보완"}
TIER_ACTION = {1: "즉시 (보고서 발행 전)", 2: "가능한 신속히", 3: "차기 감사 착수 전"}
TIER_SECTION = {
    1: "─── ★★★ 즉시 수정 필요 ───",
    2: "─── ★★ 조기 보완 필요 ───",
    3: "─── ★ 단기 보완 필요 ───",
}


# ── 규칙엔진 노트 → 스킬 ITEM 매핑 ────────────────────────────
def _importance_to_tier(importance: str) -> int:
    return {"상": 1, "중": 2, "하": 3}.get(importance, 2)


def _note_to_item(note: dict[str, Any]) -> tuple:
    """리뷰노트 dict → (tier, star, 조서, 제목, 상세, 기준, 보완지침)."""
    tier = _importance_to_tier(note.get("importance", "중"))
    section = output_formatter.format_workpaper_column(note)

    title = note.get("defect", "").strip()
    detail = output_formatter.format_issue_detail(note)

    std = output_formatter.simplify_basis(note.get("basis", ""))
    standard = std or "-"

    guide_lines: list[str] = []
    if note.get("to_be"):
        guide_lines.append(f"① {note['to_be']}")
    cases = note.get("enforcement_cases") or []
    if cases:
        primary = cases[0]
        head = f"② 감리지적사례 참고: [{primary.get('number', '')}]"
        if primary.get("subject"):
            head += f" {primary['subject']}"
        guide_lines.append(head)
        if primary.get("brief"):
            guide_lines.append(f"   - {primary['brief']}")
        others = [c for c in cases[1:] if c.get("has_number")]
        if others:
            guide_lines.append("   - 유사 지적사례: " + ", ".join(c["number"] for c in others))
    guide = "\n".join(guide_lines) or "① 관련 근거를 보완하고 문서화하십시오."

    return (tier, TIER_STAR[tier], section, title, detail, standard, guide)


# ── 표지 시트 ────────────────────────────────────────────────
def _create_cover(wb: Workbook, meta: dict) -> None:
    ws = wb.active
    ws.title = "표지"
    _no_grid(ws)
    _widths(ws, [2, 28, 58, 18])

    _mc(ws, 3, 2, 3, 4)
    c = ws.cell(row=3, column=2, value="품질관리실 감리 리뷰노트")
    c.font = _fn(bold=True, sz=22, color=COLOR["header_bg"])
    c.alignment = _al("center")
    ws.row_dimensions[3].height = 45

    _mc(ws, 4, 2, 4, 4)
    c = ws.cell(row=4, column=2, value="[ 중요사항 전용 — 중요도 우선순위별 정리 ]")
    c.font = _fn(italic=True, sz=12, color="555555")
    c.alignment = _al("center")

    for col in range(2, 5):
        ws.cell(row=5, column=col).fill = _fl(COLOR["header_bg"])
    ws.row_dimensions[5].height = 4

    info = [
        ("피감리 회사", meta.get("company", "")),
        ("결산일", meta.get("fiscal_year", "")),
        ("대상 조서", meta.get("scope", "")),
        ("감리 수행", meta.get("reviewer", "")),
        ("작성일", meta.get("date", "")),
        ("중요사항 수", meta.get("summary", "")),
    ]
    for offset, (label, val) in enumerate(info):
        row = 7 + offset
        ws.row_dimensions[row].height = 26
        _set(ws, row, 2, label, bold=True, halign="center", fill=COLOR["label_bg"])
        _mc(ws, row, 3, row, 4)
        _set(ws, row, 3, val)

    _mc(ws, 14, 2, 14, 4)
    c = ws.cell(row=14, column=2, value="[ 중요도 범례 ]")
    c.font = _fn(bold=True, sz=11, color=COLOR["header_bg"])

    legend = [
        (15, 1, "★★★  즉시 수정",
         "재무제표 왜곡표시 또는 감사기준 중대 위반 — 보고서 발행 전 반드시 해결"),
        (16, 2, "★★   조기 보완",
         "감리 지적 가능성 높은 사항 — 가능한 신속히 조서 보완 필요"),
        (17, 3, "★    단기 보완",
         "원칙적 문서화 요구 사항 — 차기 감사 착수 전까지 반드시 보완"),
    ]
    for row, tier, lbl, desc in legend:
        ws.row_dimensions[row].height = 26
        _set(ws, row, 2, lbl, bold=True, color=COLOR[f"tier{tier}_font"],
             fill=COLOR[f"tier{tier}_bg"], halign="center")
        _mc(ws, row, 3, row, 4)
        _set(ws, row, 3, desc)


# ── 우선순위 목록 시트 ────────────────────────────────────────
def _create_summary(wb: Workbook, items: list[tuple]) -> None:
    ws = wb.create_sheet("우선순위 목록")
    _no_grid(ws)
    _widths(ws, [2, 6, 14, 52, 34, 18])

    ws.row_dimensions[2].height = 32
    _mc(ws, 2, 2, 2, 7)
    c = ws.cell(row=2, column=2, value="감리 리뷰노트 — 중요사항 우선순위 목록")
    c.font = _fn(bold=True, sz=15, color=COLOR["header_bg"])
    c.alignment = _al("center")

    headers = ["No.", "우선순위", "조서", "지적사항 제목", "관련 기준", "조치 기한"]
    ws.row_dimensions[4].height = 28
    for col, h in enumerate(headers, 2):
        _set(ws, 4, col, h, bold=True, color=COLOR["header_font"],
             fill=COLOR["header_bg"], halign="center")

    prev_tier = 0
    row = 5
    for idx, item in enumerate(items, 1):
        tier, section, title, standard = item[0], item[2], item[3], item[5]
        if tier != prev_tier:
            ws.row_dimensions[row].height = 20
            _mc(ws, row, 2, row, 7)
            _set(ws, row, 2, TIER_SECTION[tier], bold=True,
                 color=COLOR[f"tier{tier}_font"], fill=COLOR[f"tier{tier}_bg"],
                 halign="center")
            row += 1
            prev_tier = tier

        ws.row_dimensions[row].height = 44
        values = [idx, TIER_LABEL[tier], section,
                  title.replace("\n", " "), standard, TIER_ACTION[tier]]
        for col, val in enumerate(values, 2):
            if col == 3:
                _set(ws, row, col, val, bold=True, color=COLOR[f"tier{tier}_font"],
                     fill=COLOR[f"tier{tier}_bg"], halign="center")
            else:
                _set(ws, row, col, val,
                     halign="center" if col in (2, 4, 7) else "left")
        row += 1


# ── 등급별 세부 시트 ──────────────────────────────────────────
def _create_detail(wb: Workbook, items: list[tuple], tier: int) -> None:
    title = f"{TIER_TITLE[tier]} ({len(items)}건)"
    ws = wb.create_sheet(title)
    _no_grid(ws)
    _widths(ws, [2, 5, 14, 52, 30, 52, 12])

    ws.row_dimensions[2].height = 34
    _mc(ws, 2, 2, 2, 7)
    c = ws.cell(row=2, column=2,
                value=f"감리 리뷰노트 세부내역 — {TIER_NAME[tier]} ({len(items)}건)")
    c.font = _fn(bold=True, sz=15, color=COLOR["header_bg"])
    c.alignment = _al("center")
    c.fill = _fl(COLOR[f"tier{tier}_bg"])

    headers = ["No.", "조서", "지적사항 제목 및 내용", "관련 기준",
               "보완지침 (감사팀 수정·보완 사항)", "조치 완료 확인"]
    ws.row_dimensions[4].height = 28
    for col, h in enumerate(headers, 2):
        _set(ws, 4, col, h, bold=True, color=COLOR["header_font"],
             fill=COLOR["header_bg"], halign="center")

    row = 5
    for idx, item in enumerate(items, 1):
        _, _, section, ttl, detail, standard, guide = item

        ws.row_dimensions[row].height = 18
        _mc(ws, row, 2, row, 7)
        _set(ws, row, 2, f"  {ttl}", bold=True, color=COLOR[f"tier{tier}_font"],
             fill=COLOR[f"tier{tier}_bg"], halign="left", wrap=False)
        row += 1

        max_lines = max(
            len(str(detail).split("\n")),
            len(str(standard).split("\n")),
            len(str(guide).split("\n")),
        )
        ws.row_dimensions[row].height = max(max_lines * 15, 90)

        body_bg = {1: "FFF8F8", 2: "FFFBF0", 3: "FFFEF8"}[tier]
        _set(ws, row, 2, idx, bold=True, halign="center", fill="F2F2F2")
        _set(ws, row, 3, section, halign="center", fill="F2F2F2")
        _set(ws, row, 4, detail, sz=9, fill=body_bg)
        _set(ws, row, 5, standard, sz=9, halign="center", fill="F0F5FF")
        _set(ws, row, 6, guide, sz=9, fill=COLOR["guide_bg"])
        _set(ws, row, 7, "□ 완료\n□ 미완\n□ N/A", sz=9, halign="center", fill="FAFAFA")
        row += 1

        _divider(ws, row)
        row += 1


def _create_focus_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """4대 중점 회계이슈 점검결과 전용 시트."""
    ws = wb.create_sheet("4대 중점 회계이슈")
    _no_grid(ws)
    _widths(ws, [2, 6, 28, 12, 48, 40])

    ws.row_dimensions[2].height = 32
    _mc(ws, 2, 2, 2, 6)
    c = ws.cell(row=2, column=2, value="4대 중점 회계이슈 점검결과")
    c.font = _fn(bold=True, sz=15, color=COLOR["header_bg"])
    c.alignment = _al("center")

    headers = ["No.", "항목", "상태", "지적사항", "근거·사유"]
    ws.row_dimensions[4].height = 26
    for col, h in enumerate(headers, 2):
        _set(ws, 4, col, h, bold=True, color=COLOR["header_font"],
             fill=COLOR["header_bg"], halign="center")

    row = 5
    for r in rows:
        ws.row_dimensions[row].height = 36
        status = r.get("status", "")
        fill = COLOR["tier1_bg"] if status in ("미비", "중점검토 필요") else "EBF7EE"
        vals = [
            r.get("issue_no", ""),
            r.get("issue_title", ""),
            status,
            r.get("defect") or "—",
            (r.get("reason") or r.get("basis") or "")[:500],
        ]
        for col, val in enumerate(vals, 2):
            _set(ws, row, col, val, fill=fill if col == 4 else None)
        row += 1


# ── 메인 ─────────────────────────────────────────────────────
def build_review_notes_excel(
    engagement: dict[str, Any],
    notes: list[dict[str, Any]],
    focus_items: list[dict[str, Any]] | None = None,
) -> bytes:
    """감리 리뷰노트 엑셀(표지·우선순위·등급별 세부)을 바이트로 생성."""
    items = sorted((_note_to_item(n) for n in notes), key=lambda it: it[0])

    n1 = sum(1 for it in items if it[0] == 1)
    n2 = sum(1 for it in items if it[0] == 2)
    n3 = sum(1 for it in items if it[0] == 3)

    year = str(engagement.get("audit_year", "")).strip()
    fiscal = f"{year}년 12월 31일" if year.isdigit() else (year or "확인 필요")
    reviewer = engagement.get("reviewer") or "품질관리실"
    meta = {
        "company": engagement.get("company_name", ""),
        "fiscal_year": fiscal,
        "scope": engagement.get("related_account", "") or "확인 필요",
        "reviewer": reviewer,
        "date": date.today().strftime("%Y년 %m월 %d일"),
        "summary": f"총 {len(items)}건 (★★★ 즉시 {n1}건 / ★★ 조기 {n2}건 / ★ 단기 {n3}건)",
    }

    wb = Workbook()
    _create_cover(wb, meta)
    _create_summary(wb, items)
    for tier in (1, 2, 3):
        tier_items = [it for it in items if it[0] == tier]
        if tier_items:
            _create_detail(wb, tier_items, tier)

    if focus_items:
        _create_focus_sheet(wb, focus_items)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
