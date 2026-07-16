"""4대 중점 체크리스트 — 시각화 xlsx 생성."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import focus_checklist_catalog as cat

_FONT = "맑은 고딕"
_ISSUE_COLORS = {
    1: "D6E4F0",
    2: "E2EFDA",
    3: "FFF2CC",
    4: "FCE4D6",
}
_GAP_COLORS = {
    "검토누락": "FFF4E5",
    "결론미비": "FDE9E9",
}


def _join(t: tuple[str, ...]) -> str:
    return "; ".join(t) if t else ""


def build_focus_checklist_workbook(*, is_listed: bool) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = cat.rows_for_listed() if is_listed else cat.rows_for_unlisted()
    src = "금융감독원(상장)" if is_listed else "한국공인회계사회(비상장)"
    year = "FY2026"

    wb = Workbook()
    # --- 개요 시트 ---
    ws0 = wb.active
    ws0.title = "개요"
    ws0["A1"] = f"4대 중점 감리 체크리스트 — {src}"
    ws0["A1"].font = Font(name=_FONT, bold=True, size=14, color="1F3864")
    ws0["A2"] = "출처: Hanul DB / 4대 중점사항 감리대상 / 2026년 보도자료 회계위반·오류 예시"
    ws0["A3"] = "용도: 감사조서 자가검토(hanul-002) · 인간 회계사 검토·보완용 living document"
    ws0["A5"] = "이슈"
    ws0["B5"] = "제목"
    ws0["C5"] = "체크 항목 수"
    seen: dict[int, tuple[str, int]] = {}
    for r in rows:
        if r.issue_no not in seen:
            seen[r.issue_no] = (r.issue_title, 0)
        t, n = seen[r.issue_no]
        seen[r.issue_no] = (t, n + 1)
    ri = 6
    for no in sorted(seen):
        title, cnt = seen[no]
        ws0.cell(ri, 1, no)
        ws0.cell(ri, 2, title)
        ws0.cell(ri, 3, cnt)
        ws0.cell(ri, 1).fill = PatternFill("solid", fgColor=_ISSUE_COLORS.get(no, "FFFFFF"))
        ri += 1
    ws0.column_dimensions["A"].width = 8
    ws0.column_dimensions["B"].width = 42
    ws0.column_dimensions["C"].width = 12

    # --- 체크리스트 시트 ---
    ws = wb.create_sheet("체크리스트")
    headers = [
        "issue_no",
        "issue_title",
        "checklist_id",
        "checklist_item",
        "violation_type",
        "case_source",
        "case_example",
        "review_gap_type",
        "materiality_note",
        "detect_all",
        "detect_any",
        "required_evidence",
        "acceptable_phrases",
        "red_flag_phrases",
        "basis",
        "to_be_if_missing",
        "to_be_if_weak",
        "golden_set_ids",
        "related_accounts",
        "sheet_keywords",
        "related_sheet_codes",
        "review_procedure",
        "ai_review_questions",
        "trigger_accounts",
        "na_condition",
        # 2026-07-16 고도화: 감리지적사례·질의회신·회계감사기준·QC 심리점검표 보강 컬럼
        "standard_paragraphs",
        "audit_standard_ref",
        "additional_case_refs",
        "qna_refs",
        "qc_checklist_ref",
    ]
    _NEW_COLS = {"standard_paragraphs", "audit_standard_ref", "additional_case_refs", "qna_refs", "qc_checklist_ref"}
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    new_hdr_fill = PatternFill("solid", fgColor="7030A0")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = new_hdr_fill if h in _NEW_COLS else hdr_fill
        c.font = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, r in enumerate(rows, 2):
        vals = [
            r.issue_no,
            r.issue_title,
            r.checklist_id,
            r.checklist_item,
            r.violation_type,
            r.case_source,
            r.case_example,
            r.review_gap_type,
            r.materiality_note,
            _join(r.detect_all),
            _join(r.detect_any),
            _join(r.required_evidence),
            _join(r.acceptable_phrases),
            _join(r.red_flag_phrases),
            r.basis,
            r.to_be_if_missing,
            r.to_be_if_weak,
            _join(r.golden_set_ids),
            _join(r.related_accounts),
            _join(r.sheet_keywords),
            _join(r.related_sheet_codes),
            r.review_procedure,
            _join(r.ai_review_questions),
            _join(r.trigger_accounts),
            r.na_condition,
            r.standard_paragraphs,
            r.audit_standard_ref,
            r.additional_case_refs,
            r.qna_refs,
            r.qc_checklist_ref,
        ]
        issue_fill = PatternFill("solid", fgColor=_ISSUE_COLORS.get(r.issue_no, "FFFFFF"))
        gap_fill = PatternFill("solid", fgColor=_GAP_COLORS.get(r.review_gap_type, "FFFFFF"))
        new_col_fill = PatternFill("solid", fgColor="F3E9FB")
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.font = Font(name=_FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if j <= 2:
                c.fill = issue_fill
            elif j == 8:
                c.fill = gap_fill
            elif headers[j - 1] in _NEW_COLS:
                c.fill = new_col_fill
        ws.row_dimensions[i].height = 48

    widths = [6, 28, 10, 28, 14, 18, 36, 10, 14, 16, 20, 18, 22, 16, 18, 28, 28, 14, 14, 16, 12, 42, 36, 14, 24,
              34, 26, 34, 26, 30]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # --- 범례 시트 ---
    ws2 = wb.create_sheet("범례")
    ws2["A1"] = "리뷰노트 유형"
    ws2["A2"] = "검토누락"
    ws2["B2"] = "조서에 해당 지적유형·검토 절차 흔적이 없음 → [4대중점·검토누락]"
    ws2["A3"] = "결론미비"
    ws2["B3"] = "검토는 있으나 감사결론·근거가 불충분 → [4대중점·결론미비]"
    ws2["A5"] = "중요성"
    ws2["B5"] = "관련 계정·거래가 중요성 미만이면 중요도 '중'으로 하향 또는 지적 생략"
    ws2["A7"] = "구체적 검토지침"
    ws2["B7"] = "review_procedure — 항목별 ①②③ 검토절차 (보도자료·감리지적사례 기반)"
    ws2["A8"] = "조서 인덱스"
    ws2["B8"] = "related_sheet_codes — FY2026 계정과목 조서 인덱스 (C, E, SAJ 등)"
    ws2["A9"] = "리뷰노트 규칙"
    ws2["B9"] = "checklist_id별 1건 · [4대중점·검토누락/결론미비] · 조서 인덱스만 표기"
    ws2["A11"] = "[신규] standard_paragraphs"
    ws2["B11"] = "K-IFRS·일반기업회계기준 구체적 조문·문단 번호 인용"
    ws2["A12"] = "[신규] audit_standard_ref"
    ws2["B12"] = "관련 회계감사기준(KSA) 조항 — 위험평가·회계추정·특수관계자·그룹감사 등"
    ws2["A13"] = "[신규] additional_case_refs"
    ws2["B13"] = "금감원·한공회 감리지적사례 추가 사례(2020~2026년, 사례코드+요약)"
    ws2["A14"] = "[신규] qna_refs"
    ws2["B14"] = "한국회계기준원 질의회신 인용(번호+주제)"
    ws2["A15"] = "[신규] qc_checklist_ref"
    ws2["B15"] = "한울 사전/사후심리 체크리스트(QC) 연계 항목·미흡판정기준"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deploy_to_hanul_db() -> tuple[str, str]:
    """Hanul DB 자가검토_지침_템플릿에 배포 + data/templates 미러."""
    import knowledge_base as kb
    import review_guidelines as rg

    root = kb.SOURCE_DIR
    if isinstance(root, str):
        root = Path(root)
    target = root / rg.GUIDELINES_DB_SUBDIR
    target.mkdir(parents=True, exist_ok=True)
    listed_path = target / rg.TEMPLATE_FILES["focus_listed"]
    unlisted_path = target / rg.TEMPLATE_FILES["focus_unlisted"]
    listed_bytes = build_focus_checklist_workbook(is_listed=True)
    unlisted_bytes = build_focus_checklist_workbook(is_listed=False)
    listed_path.write_bytes(listed_bytes)
    unlisted_path.write_bytes(unlisted_bytes)

    local = Path(__file__).resolve().parent / "data" / "templates"
    local.mkdir(parents=True, exist_ok=True)
    (local / rg.TEMPLATE_FILES["focus_listed"]).write_bytes(listed_bytes)
    (local / rg.TEMPLATE_FILES["focus_unlisted"]).write_bytes(unlisted_bytes)
    return str(listed_path), str(unlisted_path)


if __name__ == "__main__":
    a, b = deploy_to_hanul_db()
    print(a)
    print(b)
