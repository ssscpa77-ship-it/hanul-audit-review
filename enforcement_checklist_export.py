"""계정·조서인덱스별 감리지적 체크리스트 xlsx 생성·Hanul DB 배포."""

from __future__ import annotations

from io import BytesIO

import enforcement_checklist_catalog as cat
import enforcement_case_learner as ecl

_FONT = "맑은 고딕"


def _join(t: tuple[str, ...] | str) -> str:
    if isinstance(t, str):
        return t
    return "; ".join(t) if t else ""


def build_enforcement_checklist_workbook(*, is_listed: bool) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    variant = "ifrs_listed" if is_listed else "non_listed"
    data = ecl.learn_and_build(variant=variant)
    rows: list[ecl.EnforcementCheckRow] = data["rows"]
    summaries: list[ecl.AccountKeywordSummary] = data["summaries"]
    src = "금융감독원·한공회 감리지적사례 (Hanul DB 학습)"
    listing = "상장·K-IFRS" if is_listed else "비상장"

    wb = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")

    # --- 개요 ---
    ws0 = wb.active
    ws0.title = "개요"
    ws0["A1"] = f"계정·조서인덱스별 감리지적 체크리스트 — {listing} FY2026"
    ws0["A1"].font = Font(name=_FONT, bold=True, size=14, color="1F3864")
    ws0["A2"] = f"출처: {src}"
    ws0["A3"] = "매핑: 4000_계정별 실증절차 조서 인덱스 (A, C, D, E, SAJ …)"
    ws0["A4"] = f"학습 사례: {data['total_cases']}건 → 체크리스트 {len(rows)}항목 · 조서인덱스 {len({r.sheet_code for r in rows})}개"
    ws0["A6"] = "용도: 감사조서 자가검토 — 감리지적 유형·맥락·점검포인트 기반 절차 흠결 탐지"
    ws0["A7"] = "※ case_context·audit_focus 열에 사례 맥락·구체 점검항목이 상세 기재됨"
    ws0["A8"] = "조서인덱스"
    ws0["B8"] = "계정과목"
    ws0["C8"] = "지적사례 수"
    ws0["D8"] = "체크 항목 수"
    for c in range(1, 5):
        ws0.cell(8, c).fill = hdr_fill
        ws0.cell(8, c).font = Font(name=_FONT, bold=True, color="FFFFFF")
    by_code: dict[str, tuple[str, int, int]] = {}
    for r in rows:
        if r.sheet_code not in by_code:
            by_code[r.sheet_code] = (r.sheet_label, 0, 0)
        lbl, cc, ci = by_code[r.sheet_code]
        by_code[r.sheet_code] = (lbl, cc + r.case_count, ci + 1)
    ri = 9
    for code in sorted(by_code.keys(), key=lambda x: (len(x), x)):
        lbl, cc, ci = by_code[code]
        ws0.cell(ri, 1, code)
        ws0.cell(ri, 2, lbl)
        ws0.cell(ri, 3, cc)
        ws0.cell(ri, 4, ci)
        ri += 1
    ws0.column_dimensions["A"].width = 12
    ws0.column_dimensions["B"].width = 42
    ws0.column_dimensions["C"].width = 12
    ws0.column_dimensions["D"].width = 12

    # --- 키워드요약 ---
    ws_kw = wb.create_sheet("키워드요약")
    kw_headers = [
        "sheet_code", "sheet_label", "canonical_account",
        "case_count", "violation_types", "key_keywords", "case_numbers", "case_narratives",
    ]
    for j, h in enumerate(kw_headers, 1):
        c = ws_kw.cell(1, j, h)
        c.fill = hdr_fill
        c.font = Font(name=_FONT, bold=True, color="FFFFFF")
        c.border = border
    for i, s in enumerate(summaries, 2):
        vals = [
            s.sheet_code, s.sheet_label, s.account, s.case_count,
            _join(s.violation_types), _join(s.key_keywords), _join(s.case_numbers),
            s.case_narratives,
        ]
        for j, v in enumerate(vals, 1):
            ws_kw.cell(i, j, v).font = Font(name=_FONT, size=10)
            ws_kw.cell(i, j).border = border
    for idx, w in enumerate([10, 36, 16, 10, 28, 36, 28, 72], 1):
        ws_kw.column_dimensions[get_column_letter(idx)].width = w
    ws_kw.freeze_panes = "A2"

    # --- 체크리스트 ---
    ws = wb.create_sheet("체크리스트")
    headers = [
        "sheet_code", "sheet_label", "canonical_account", "checklist_id",
        "checklist_item", "violation_type", "procedure_gap",
        "case_source", "case_count", "case_numbers", "case_examples",
        "case_context", "audit_focus",
        "key_keywords", "review_gap_type",
        "detect_any", "required_evidence", "acceptable_phrases", "red_flag_phrases",
        "review_procedure", "to_be",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = hdr_fill
        c.font = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    code_colors = [
        "D6E4F0", "E2EFDA", "FFF2CC", "FCE4D6", "EAD1F5", "E8E8E8",
    ]
    color_map: dict[str, str] = {}
    for i, r in enumerate(rows):
        if r.sheet_code not in color_map:
            color_map[r.sheet_code] = code_colors[len(color_map) % len(code_colors)]

    for i, r in enumerate(rows, 2):
        vals = [
            r.sheet_code, r.sheet_label, r.canonical_account, r.checklist_id,
            r.checklist_item, r.violation_type, r.procedure_gap,
            r.case_source, r.case_count, r.case_numbers, r.case_examples,
            r.case_context, r.audit_focus,
            r.key_keywords, r.review_gap_type,
            _join(r.detect_any), _join(r.required_evidence),
            _join(r.acceptable_phrases), _join(r.red_flag_phrases),
            r.review_procedure, r.to_be,
        ]
        fill = PatternFill("solid", fgColor=color_map.get(r.sheet_code, "FFFFFF"))
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.font = Font(name=_FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
            if j <= 3:
                c.fill = fill
        ws.row_dimensions[i].height = 72

    widths = [8, 32, 14, 10, 40, 12, 48, 14, 8, 24, 56, 56, 40, 28, 10, 28, 18, 22, 18, 48, 44]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # --- 범례 ---
    ws2 = wb.create_sheet("범례")
    ws2["A1"] = "감리지적 체크리스트 사용법"
    ws2["A2"] = "sheet_code"
    ws2["B2"] = "4000 계정별 실증절차 조서 인덱스 (한울 FY2026 목차)"
    ws2["A3"] = "procedure_gap"
    ws2["B3"] = "감리에서 지적된 절차 흠결 유형 요약 — 조서에서 동일 유형 누락 탐지"
    ws2["A4"] = "case_context"
    ws2["B4"] = "감리지적 사례 맥락·지적 배경 요약 (리뷰노트 reason·근거 연결)"
    ws2["A5"] = "audit_focus"
    ws2["B5"] = "조서에서 확인할 구체 점검 항목"
    ws2["A6"] = "detect_any"
    ws2["B6"] = "조서 본문에 해당 키워드가 있으면 이 항목 검토 대상"
    ws2["A7"] = "required_evidence"
    ws2["B7"] = "필수 검토·문서화 흔적 — 없으면 [감리지적·검토누락] 노트"
    ws2["A8"] = "업데이트"
    ws2["B8"] = "인간 회계사·AI가 Hanul DB 신규 감리사례 반영 시 재학습·재배포"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deploy_to_hanul_db() -> tuple[str, str]:
    """Hanul DB 자가검토_지침_템플릿에 배포."""
    import knowledge_base as kb
    import review_guidelines as rg
    from pathlib import Path

    # 캐시 초기화 — 상세 요약 반영
    try:
        import enforcement_checklist_catalog as ecc
        ecc.rows_for_listed.cache_clear()
        ecc.rows_for_unlisted.cache_clear()
        ecc.summaries_for_listed.cache_clear()
        ecc.summaries_for_unlisted.cache_clear()
    except Exception:  # noqa: BLE001
        pass
    try:
        import guidelines_loader as gl
        gl.load_enforcement_checklist_from_db.cache_clear()
    except Exception:  # noqa: BLE001
        pass

    root = Path(kb.SOURCE_DIR)
    target = root / rg.GUIDELINES_DB_SUBDIR
    target.mkdir(parents=True, exist_ok=True)
    listed_path = target / rg.TEMPLATE_FILES["enforcement_listed"]
    unlisted_path = target / rg.TEMPLATE_FILES["enforcement_unlisted"]
    listed_path.write_bytes(build_enforcement_checklist_workbook(is_listed=True))
    unlisted_path.write_bytes(build_enforcement_checklist_workbook(is_listed=False))
    return str(listed_path), str(unlisted_path)


if __name__ == "__main__":
    a, b = deploy_to_hanul_db()
    print("배포 완료:")
    print(a)
    print(b)
    print("상장:", cat.stats(is_listed=True))
    print("비상장:", cat.stats(is_listed=False))
